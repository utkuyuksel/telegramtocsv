import os
import csv
import json
import random
import asyncio
import traceback
import zipfile
from datetime import datetime

from pyrogram import Client
from pyrogram.errors import FloodWait

import config

DOWNLOADS_DIR = config.DOWNLOADS_DIR
SESSION_FILE = config.SESSION_FILE
DEAD_SESSION_FILE = config.DEAD_SESSION_FILE


def _csv_safe(value):
    """Neutralize CSV/formula injection so the file is safe in Excel/Sheets.
    A leading =, +, -, @ (or tab/CR) can run as a formula; prefix it with a quote."""
    s = "" if value is None else str(value)
    if s[:1] in ("=", "+", "-", "@", "\t", "\r"):
        s = "'" + s
    return s


EXPORT_HEADER = ["Message ID", "Date", "Content", "Views", "Link"]
EXCEL_MAX_ROWS = 1_048_576  # Excel's hard per-sheet row limit (incl. header)


class _ExportWriter:
    """Streams message rows to CSV or XLSX.

    Both formats neutralize spreadsheet formula injection: CSV is UTF-8 with BOM
    (Excel-friendly) and prefixes a leading =,+,-,@ with a quote; XLSX forces the
    Content cell to a text type so Excel never parses a leading '=' as a formula
    (openpyxl would otherwise store it as a real formula). XLSX rolls onto a new
    sheet at Excel's row limit so arbitrarily large channels still open.
    """

    def __init__(self, fmt, path):
        self.fmt = "xlsx" if str(fmt).lower() == "xlsx" else "csv"
        self.path = path
        if self.fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.cell import WriteOnlyCell

            self._WOCell = WriteOnlyCell
            self._wb = Workbook(write_only=True)
            self._ws = self._wb.create_sheet("Messages")
            self._ws.append(EXPORT_HEADER)
            self._rows = 1
            self._sheet = 1
        else:
            self._f = open(path, "w", encoding="utf-8-sig", newline="")
            self._w = csv.writer(self._f)
            self._w.writerow(EXPORT_HEADER)

    def _text_cell(self, value):
        # Force a text cell so Excel never runs a leading '='/'+'/'-'/'@' as a formula.
        cell = self._WOCell(self._ws, value="" if value is None else str(value))
        cell.data_type = "s"
        return cell

    def write(self, msg_id, date, content, views, link):
        if self.fmt == "xlsx":
            if self._rows >= EXCEL_MAX_ROWS:
                self._sheet += 1
                self._ws = self._wb.create_sheet(f"Messages {self._sheet}")
                self._ws.append(EXPORT_HEADER)
                self._rows = 1
            self._ws.append([msg_id, date, self._text_cell(content), views, link])
            self._rows += 1
        else:
            self._w.writerow([msg_id, date, _csv_safe(content), views, link])

    def close(self):
        if self.fmt == "xlsx":
            self._wb.save(self.path)
        else:
            self._f.close()


def load_sessions():
    try:
        with open(SESSION_FILE, "r") as f:
            return json.load(f)
    except Exception:
        return {}


SESSION_STRINGS = load_sessions()


def remove_bad_session(worker_name, error_reason="Unknown"):
    global SESSION_STRINGS
    try:
        if worker_name in SESSION_STRINGS:
            print(f"[SYSTEM] Dead session removed: {worker_name}")
            dead_entry = {
                "session_string": SESSION_STRINGS[worker_name],
                "died_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                "reason": str(error_reason),
            }
            dead_sessions = {}
            if os.path.exists(DEAD_SESSION_FILE):
                try:
                    with open(DEAD_SESSION_FILE, "r") as f:
                        dead_sessions = json.load(f)
                except Exception:
                    pass

            dead_sessions[worker_name] = dead_entry
            with open(DEAD_SESSION_FILE, "w") as f:
                json.dump(dead_sessions, f, indent=2)

            del SESSION_STRINGS[worker_name]
            with open(SESSION_FILE, "w") as f:
                json.dump(SESSION_STRINGS, f, indent=2)
    except Exception:
        pass


class SessionManager:
    def __init__(self, worker_prefix=""):
        self.session_items = list(SESSION_STRINGS.items())
        if self.session_items:
            random.shuffle(self.session_items)

        self.worker_queue = self.session_items.copy()
        self.active_clients = {}
        self.worker_prefix = worker_prefix

    async def get_next_worker(self):
        if not self.worker_queue:
            global SESSION_STRINGS
            SESSION_STRINGS = load_sessions()
            if not SESSION_STRINGS:
                raise Exception("No active workers available! Please add accounts.")
            self.session_items = list(SESSION_STRINGS.items())
            self.worker_queue = self.session_items.copy()
            random.shuffle(self.worker_queue)

        worker_name, session_string = self.worker_queue.pop(0)

        if worker_name in self.active_clients:
            return self.active_clients[worker_name], worker_name

        return await self._start_worker(worker_name, session_string)

    async def _start_worker(self, worker_name, session_string):
        try:
            app = Client(
                name=f"{self.worker_prefix}_{worker_name}",
                api_id=config.API_ID,
                api_hash=config.API_HASH,
                session_string=session_string,
                in_memory=True,
                sleep_threshold=0,
                no_updates=True,
            )
            await app.start()
            self.active_clients[worker_name] = app
            return app, worker_name
        except Exception as e:
            remove_bad_session(worker_name, error_reason=e)
            return await self.get_next_worker()

    async def rotate_worker(self, current_app, current_worker):
        try:
            await current_app.stop()
        except Exception:
            pass
        if current_worker in self.active_clients:
            del self.active_clients[current_worker]
        return await self.get_next_worker()

    async def cleanup(self):
        for _, app in self.active_clients.items():
            try:
                await app.stop()
            except Exception:
                pass
        self.active_clients.clear()


async def process_scraping(
    channel_link,
    order_token,
    limit=None,
    include_media=False,
    fmt="csv",
    progress_callback=None,
    worker_callback=None,
):
    """
    Scrape a Telegram channel into a zipped CSV.

    limit: hard cap on messages (None = unlimited; free tier passes ~100,
           paid tier passes PAID_MAX_MESSAGES).
    worker_callback: called with (worker_name) when a worker is selected, so
                     the caller can persist which worker handled this order.
    Returns: (zip_path, processed_count, total_count) on success,
             or (None, 0, 0) on failure.
             `total_count` is the channel's actual message count (so callers
             can detect when a paid cap truncated a large channel).
    """
    try:
        manager = SessionManager(worker_prefix=f"[{order_token[:5]}]")
    except Exception:
        return None, 0, 0

    app = None
    processed = 0
    total_count = 0

    try:
        channel_name = channel_link.split("/")[-1].replace("@", "").strip()

        if progress_callback:
            progress_callback(10, "Connecting to Telegram...")
        app, current_worker = await manager.get_next_worker()
        if worker_callback:
            worker_callback(current_worker)

        try:
            chat = await app.get_chat(channel_name)
            total_count = await app.get_chat_history_count(channel_name)
        except Exception as e:
            print(f"Channel not found: {e}")
            return None, 0, 0

        if total_count == 0:
            return None, 0, 0

        target_count = total_count if limit is None else min(limit, total_count)

        os.makedirs(DOWNLOADS_DIR, exist_ok=True)
        fmt = "xlsx" if str(fmt).lower() == "xlsx" else "csv"
        ext = "xlsx" if fmt == "xlsx" else "csv"
        base_filename = f"{channel_name}_{order_token}"
        data_path = os.path.join(DOWNLOADS_DIR, f"{base_filename}.{ext}")

        writer = _ExportWriter(fmt, data_path)
        offset_id = 0
        try:
            while processed < target_count:
                try:
                    messages = []
                    async for msg in app.get_chat_history(
                        channel_name, limit=100, offset_id=offset_id
                    ):
                        messages.append(msg)

                    if not messages:
                        break

                    for msg in messages:
                        if processed >= target_count:
                            break
                        processed += 1
                        try:
                            text = msg.text or msg.caption or "[Media/File]"
                            link = f"https://t.me/{channel_name}/{msg.id}"
                            date = (
                                msg.date.strftime("%Y-%m-%d %H:%M:%S")
                                if msg.date
                                else ""
                            )
                            views = msg.views or 0
                            writer.write(msg.id, date, text, views, link)
                            offset_id = msg.id
                        except Exception:
                            continue

                        if progress_callback and processed % 50 == 0:
                            percent = 10 + int((processed / target_count) * 80)
                            progress_callback(
                                percent,
                                f"Scraping: {processed}/{target_count} messages...",
                            )

                except FloodWait as e:
                    print(f"FloodWait: {e.value}s. Rotating worker...")
                    app, current_worker = await manager.rotate_worker(
                        app, current_worker
                    )
                    if worker_callback:
                        worker_callback(current_worker)
                except Exception:
                    app, current_worker = await manager.rotate_worker(
                        app, current_worker
                    )
                    if worker_callback:
                        worker_callback(current_worker)
        finally:
            writer.close()

        if fmt == "xlsx":
            # .xlsx is already a zipped container — serve it directly.
            result_path = data_path
        else:
            if progress_callback:
                progress_callback(95, "Compressing file...")
            result_path = os.path.join(DOWNLOADS_DIR, f"{base_filename}.zip")
            with zipfile.ZipFile(result_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(data_path, arcname=f"{channel_name}.csv")
            try:
                os.remove(data_path)
            except Exception:
                pass

        if progress_callback:
            progress_callback(100, "Completed!")
        return result_path, processed, total_count

    except Exception:
        traceback.print_exc()
        return None, processed, total_count
    finally:
        await manager.cleanup()


async def get_channel_count(channel_link):
    """
    Validate a public channel and return its message count WITHOUT scraping.
    Returns (count, None) on success, or (None, error_message) on failure.
    Cheap enough to call before payment to produce a by-size quote.
    """
    try:
        manager = SessionManager(worker_prefix="[quote]")
    except Exception:
        return None, "Service is busy. Please try again shortly."
    try:
        channel_name = channel_link.split("/")[-1].replace("@", "").strip()
        if not channel_name:
            return None, "Please enter a valid public channel link."
        try:
            app, _worker = await manager.get_next_worker()
        except Exception:
            return None, "Service is busy. Please try again shortly."
        try:
            await app.get_chat(channel_name)
            count = await app.get_chat_history_count(channel_name)
        except Exception:
            return None, "Channel not found. It must be public and non-empty."
        if not count:
            return None, "This channel appears to be empty."
        return int(count), None
    finally:
        await manager.cleanup()
